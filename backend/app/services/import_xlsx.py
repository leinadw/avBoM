"""Import existing .xlsx project files — reads the SMW Equipment List template format."""
import uuid
from decimal import Decimal, InvalidOperation
from typing import Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.project import Project, ProjectMember
from app.models.system import System, SystemSection, SystemItem, SystemType
from app.models.equipment import Equipment
from app.models.user import User

EXCLUDED_SHEETS = {
    "Summary", "SYSTEM_TEMPLATE_LOOKUP", "DATA_HOLD", "PROJECT_EQUIPMENT_LIST",
    "PROJECT_SETTINGS", "INSTRUCTIONS", "Issuances", "Revision List",
    "_TEMP", "Equipment Report", "DWG Report", "Cutsheet Report", "Equipment Cost",
}

# Row color hex codes from VBA (as openpyxl RGB)
SECTION_HEADER_COLOR = "FFD9B600"  # approx 14270668 decimal = #D9B600
END_MARKER_COLORS = {"FFDA9694", "FFC6EFCE"}  # approx 14277081, 13288897


def _safe_decimal(val) -> Decimal:
    try:
        if val is None or val == "":
            return Decimal("0")
        return Decimal(str(val))
    except InvalidOperation:
        return Decimal("0")


def _cell_bg(cell) -> Optional[str]:
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            return fill.fgColor.rgb
    except Exception:
        pass
    return None


async def import_xlsx(
    file_bytes: bytes,
    owner: User,
    db: AsyncSession,
) -> Project:
    wb = load_workbook(filename=__import__("io").BytesIO(file_bytes), data_only=True)

    # --- Read PROJECT_SETTINGS ---
    project_name = "[Imported Project]"
    discount_pct = Decimal("0")
    contingency_pct = Decimal("0")
    rounding_variable = 0

    engineering_mult = Decimal("0.02")
    pm_mult = Decimal("0.02")
    preinstall_mult = Decimal("0.02")
    installation_mult = Decimal("0.08")
    programming_mult = Decimal("0.08")
    tax_mult = Decimal("0.08275")
    ga_mult = Decimal("0.10")

    if "PROJECT_SETTINGS" in wb.sheetnames:
        ps = wb["PROJECT_SETTINGS"]
        project_name = ps["A3"].value or "[Imported Project]"
        rounding_variable = int(_safe_decimal(ps["F3"].value))
        discount_pct = _safe_decimal(ps["H3"].value)
        contingency_pct = _safe_decimal(ps["J3"].value)
        # Non-equip multipliers: B col = label, C col = mult, starting row 2
        mult_map = {}
        for row in ps.iter_rows(min_row=2, max_row=12, values_only=True):
            if row[1] and row[2]:
                mult_map[str(row[1]).strip().upper()] = _safe_decimal(row[2])
        if mult_map:
            engineering_mult = mult_map.get("ENGINEERING", engineering_mult)
            pm_mult = mult_map.get("PROJECT MANAGEMENT", pm_mult)
            installation_mult = mult_map.get("INSTALLATION", installation_mult)
            programming_mult = mult_map.get("PROGRAMMING", programming_mult)
            tax_mult = mult_map.get("TAX", tax_mult)
            ga_mult = mult_map.get("G&A", ga_mult)

    # --- Build Equipment lookup from PROJECT_EQUIPMENT_LIST ---
    equip_lookup: dict[str, dict] = {}
    if "PROJECT_EQUIPMENT_LIST" in wb.sheetnames:
        el = wb["PROJECT_EQUIPMENT_LIST"]
        for row in el.iter_rows(min_row=2, values_only=True):
            item_id = str(row[0]).strip() if row[0] else None
            if not item_id or item_id == "//":
                continue
            equip_lookup[item_id] = {
                "mfr": str(row[1] or "").strip(),
                "model": str(row[2] or "").strip(),
                "description": str(row[3] or "").strip(),
                "notes": str(row[4] or "").strip(),
                "msrp": _safe_decimal(row[5]),
                "multiplier": _safe_decimal(row[6]) or Decimal("1.0"),
            }

    # --- Create Project ---
    project = Project(
        name=str(project_name),
        created_by_id=owner.id,
        rounding_variable=rounding_variable,
        discount_pct=discount_pct,
        contingency_pct=contingency_pct,
        engineering_mult=engineering_mult,
        pm_mult=pm_mult,
        preinstall_mult=preinstall_mult,
        installation_mult=installation_mult,
        programming_mult=programming_mult,
        tax_mult=tax_mult,
        ga_mult=ga_mult,
    )
    db.add(project)
    await db.flush()

    # Add owner as member
    db.add(ProjectMember(project_id=project.id, user_id=owner.id, role="owner"))

    # --- Import system sheets ---
    display_order = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDED_SHEETS:
            continue
        ws = wb[sheet_name]

        # Detect room info from row 2
        c2_val = ws["C2"].value
        d2_val = ws["D2"].value
        if c2_val == "Room Numbers":
            system_type = SystemType.room_numbers
            room_info = str(d2_val or "")
        elif c2_val == "System Count":
            system_type = SystemType.system_count
            room_info = str(d2_val or "1")
        else:
            system_type = SystemType.room_numbers
            room_info = ""

        system = System(
            project_id=project.id,
            name=sheet_name,
            system_type=system_type,
            room_info=room_info,
            display_order=display_order,
        )
        db.add(system)
        await db.flush()
        display_order += 1

        # Parse rows starting at row 5 (header) / 6 (data)
        current_section: Optional[SystemSection] = None
        section_order = 0
        item_order = 0

        for row_idx in range(5, ws.max_row + 1):
            a_cell = ws.cell(row=row_idx, column=1)
            b_cell = ws.cell(row=row_idx, column=2)
            f_cell = ws.cell(row=row_idx, column=6)
            g_cell = ws.cell(row=row_idx, column=7)
            j_cell = ws.cell(row=row_idx, column=10)
            k_cell = ws.cell(row=row_idx, column=11)

            bg = _cell_bg(a_cell)

            # End marker row
            if bg in END_MARKER_COLORS:
                break

            a_val = a_cell.value
            b_val = b_cell.value

            # Section header row (colored but not end marker)
            if bg and bg not in END_MARKER_COLORS:
                section = SystemSection(
                    system_id=system.id,
                    name=str(a_val or ""),
                    display_order=section_order,
                )
                db.add(section)
                await db.flush()
                current_section = section
                section_order += 1

                # Add as section header item too
                item = SystemItem(
                    system_id=system.id,
                    section_id=section.id,
                    display_order=item_order,
                    is_section_header=True,
                    note_text=str(a_val or ""),
                )
                db.add(item)
                item_order += 1
                continue

            # Note row (no item_id, has text in col B or A)
            if not a_val or str(a_val).strip() == "":
                if b_val:
                    item = SystemItem(
                        system_id=system.id,
                        section_id=current_section.id if current_section else None,
                        display_order=item_order,
                        is_note_row=True,
                        note_text=str(b_val),
                    )
                    db.add(item)
                    item_order += 1
                continue

            item_id_str = str(a_val).strip()
            if item_id_str == "//":
                break

            # Regular equipment row
            equip_data = equip_lookup.get(item_id_str, {})
            msrp = _safe_decimal(j_cell.value) or _safe_decimal(equip_data.get("msrp", 0))
            mult = _safe_decimal(k_cell.value) or _safe_decimal(equip_data.get("multiplier", 1)) or Decimal("1")
            qty = _safe_decimal(f_cell.value)

            # Check OFCI
            g_val = g_cell.value
            is_ofci = False
            ofci_type_val = None
            if g_val and isinstance(g_val, str) and g_val.strip().upper() in ("OFE", "OFCI", "OFOI"):
                is_ofci = True
                ofci_type_val = g_val.strip().upper()
            elif g_val and isinstance(g_val, str) and not g_val.replace(".", "").replace("-", "").isnumeric():
                is_ofci = True
                ofci_type_val = g_val.strip()

            item = SystemItem(
                system_id=system.id,
                section_id=current_section.id if current_section else None,
                display_order=item_order,
                item_id=item_id_str,
                description=str(equip_data.get("description", b_val or "")),
                notes=str(equip_data.get("notes", "")),
                qty_per_room=qty,
                msrp=msrp,
                multiplier=mult,
                is_ofci=is_ofci,
                ofci_type=ofci_type_val,
            )
            db.add(item)
            item_order += 1

    await db.commit()
    await db.refresh(project)
    return project
