"""
Export service — generates Publish BoM (Appendix A) and Publish Estimate (EoPC) Excel files.
Mirrors the VBA PubBoM / pubEST / CleanSheet logic.
"""
import io
from decimal import Decimal
from typing import List, Optional
from datetime import date
import uuid

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from app.models.project import Project
from app.models.system import System, SystemItem, SystemSection
from app.services.summary import compute_system_summary, _round_value


# ─── Style constants ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=10)
OFCI_FILL = PatternFill("solid", fgColor="FFE699")
CHANGED_UP_FILL = PatternFill("solid", fgColor="F4CCCC")
CHANGED_DOWN_FILL = PatternFill("solid", fgColor="D9EAD3")
THIN = Side(style="thin")
MEDIUM = Side(style="medium")
CURRENCY_FMT = '#,##0.00'
PCT_FMT = '0.00%'


def _col(n: int) -> str:
    return get_column_letter(n)


def _apply_border(cell, top=None, bottom=None, left=None, right=None):
    cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def _write_system_sheet(
    wb: Workbook,
    system: System,
    project: Project,
    issuance_name: str,
    is_bom: bool,
    include_notes: bool,
    include_cost: bool,
    include_labor_breakout: bool,
) -> None:
    ws = wb.create_sheet(title=system.name[:31])

    # ── Row 1: Project name + issuance date ──────────────────────────────────
    ws["A1"] = project.name
    ws["A1"].font = Font(bold=True, size=12)
    ws["I1"] = date.today().isoformat()

    # ── Row 2: System name + room info ───────────────────────────────────────
    ws["A2"] = system.name
    ws["A2"].font = Font(bold=True, size=11)
    ws["C2"] = system.system_type.value.replace("_", " ").title()
    ws["D2"] = system.room_info or ""

    # ── Row 3: Issuance ──────────────────────────────────────────────────────
    ws["A3"] = issuance_name

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    headers = ["ITEM", "DESCRIPTION", "MODEL", "NOTES", "QTY", "UNIT COST", "EXTENDED COST"]
    if not is_bom:
        headers = ["ITEM", "DESCRIPTION", "MODEL", "NOTES", "QTY", "MSRP", "MULT", "UNIT COST", "EXTENDED COST"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 5
    section_start_rows: List[int] = []
    section_end_rows: List[int] = []
    current_section_start = row

    items = sorted(system.items, key=lambda i: i.display_order)

    for item in items:
        if item.is_section_header:
            section_start_rows.append(row)
            ws.cell(row=row, column=1, value=item.note_text or "").fill = SECTION_FILL
            ws.merge_cells(f"A{row}:I{row}")
            ws.cell(row=row, column=1).font = SECTION_FONT
            row += 1
            continue

        if item.is_note_row:
            ws.cell(row=row, column=2, value=item.note_text or "")
            if item.is_bold_note:
                ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1
            continue

        # Regular equipment row
        ws.cell(row=row, column=1, value=item.item_id or "")
        ws.cell(row=row, column=2, value=item.description or "")
        ws.cell(row=row, column=3, value="")  # model col (from equip lookup)
        if not include_notes:
            ws.cell(row=row, column=4, value="")
        else:
            ws.cell(row=row, column=4, value=item.notes or "")

        qty = float(item.qty_per_room)
        ws.cell(row=row, column=5, value=qty)

        if item.is_ofci:
            ofci_label = item.ofci_type or "OFCI"
            ws.cell(row=row, column=6, value=ofci_label)
            ws.cell(row=row, column=7, value=ofci_label)
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = OFCI_FILL
        else:
            unit_cost = float(item.msrp * item.multiplier)
            extended = unit_cost * qty

            if is_bom:
                ws.cell(row=row, column=6, value=unit_cost).number_format = CURRENCY_FMT
                ws.cell(row=row, column=7, value=extended).number_format = CURRENCY_FMT
            else:
                ws.cell(row=row, column=6, value=float(item.msrp)).number_format = CURRENCY_FMT
                ws.cell(row=row, column=7, value=float(item.multiplier)).number_format = '0.000'
                ws.cell(row=row, column=8, value=unit_cost).number_format = CURRENCY_FMT
                ws.cell(row=row, column=9, value=extended).number_format = CURRENCY_FMT

            # Change highlighting
            if item.change_status == "increased":
                for c in range(1, 8):
                    ws.cell(row=row, column=c).fill = CHANGED_UP_FILL
            elif item.change_status == "decreased":
                for c in range(1, 8):
                    ws.cell(row=row, column=c).fill = CHANGED_DOWN_FILL

        row += 1

    # ── Cost summary section ──────────────────────────────────────────────────
    summary = compute_system_summary(system, project)
    eq_col = 7 if is_bom else 9  # extended cost column

    ws.cell(row=row, column=1, value="//").fill = PatternFill("solid", fgColor="C6EFCE")
    row += 1

    # Sub-totals
    def _sum_label(label, value, r):
        ws.cell(row=r, column=eq_col - 1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=eq_col, value=float(value)).number_format = CURRENCY_FMT

    _sum_label("TOTAL EQUIPMENT COST SUBTOTAL", summary.equipment_subtotal, row); row += 1

    if not is_bom:
        _sum_label("DISCOUNT FROM MSRP", -summary.discount_amount, row); row += 1
        _sum_label("DISCOUNTED EQUIPMENT COST SUBTOTAL", summary.discounted_equipment, row); row += 1
        _sum_label("TOTAL NON-EQUIPMENT COST SUBTOTAL", summary.non_equipment_subtotal, row); row += 1

        if include_labor_breakout:
            ne_mults = [
                (project.engineering_label, project.engineering_mult),
                (project.pm_label, project.pm_mult),
                (project.preinstall_label, project.preinstall_mult),
                (project.installation_label, project.installation_mult),
                (project.programming_label, project.programming_mult),
                (project.tax_label, project.tax_mult),
                (project.ga_label, project.ga_mult),
            ]
            for label, mult in ne_mults:
                val = _round_value(summary.discounted_equipment * mult, project.rounding_variable)
                ws.cell(row=row, column=eq_col - 2, value=label)
                ws.cell(row=row, column=eq_col - 1, value=float(mult)).number_format = PCT_FMT
                ws.cell(row=row, column=eq_col, value=float(val)).number_format = CURRENCY_FMT
                row += 1

        _sum_label("CONTINGENCY PERCENTAGE", summary.contingency_pct, row); row += 1
        _sum_label("CONTINGENCY", summary.contingency_amount, row); row += 1

    _sum_label("TOTAL INSTALLED COST", summary.system_subtotal, row); row += 1

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14


def _write_summary_sheet(wb: Workbook, project: Project, systems: List[System], issuance_name: str) -> None:
    ws = wb.create_sheet(title="Summary", index=0)
    ws["A1"] = project.name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = issuance_name

    headers = ["ID", "DESCRIPTION", "ROOM #'s", "EQUIPMENT", "DISCOUNT", "DISCOUNTED EQUIP",
               "NON-EQUIPMENT", "CONTINGENCY %", "CONTINGENCY", "SYSTEM SUBTOTAL", "QTY", "SYSTEM EXTENDED"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    totals = {
        "equipment": Decimal("0"), "discount": Decimal("0"), "discounted": Decimal("0"),
        "non_equip": Decimal("0"), "contingency": Decimal("0"), "subtotal": Decimal("0"), "extended": Decimal("0"),
    }

    for row_idx, system in enumerate(systems, 5):
        s = compute_system_summary(system, project)
        row_data = [
            system.name, system.name, system.room_info or "",
            float(s.equipment_subtotal), float(s.discount_amount), float(s.discounted_equipment),
            float(s.non_equipment_subtotal), float(s.contingency_pct), float(s.contingency_amount),
            float(s.system_subtotal), s.room_count, float(s.system_extended),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in (4, 5, 6, 7, 9, 10, 12):
                cell.number_format = CURRENCY_FMT
            elif col_idx == 8:
                cell.number_format = PCT_FMT

        totals["equipment"] += s.equipment_subtotal
        totals["discount"] += s.discount_amount
        totals["discounted"] += s.discounted_equipment
        totals["non_equip"] += s.non_equipment_subtotal
        totals["contingency"] += s.contingency_amount
        totals["subtotal"] += s.system_subtotal
        totals["extended"] += s.system_extended

    # Totals row
    tr = len(systems) + 5
    ws.cell(row=tr + 2, column=3, value="TOTAL EQUIPMENT COST SUBTOTAL").font = Font(bold=True)
    ws.cell(row=tr + 2, column=4, value=float(totals["equipment"])).number_format = CURRENCY_FMT
    ws.cell(row=tr + 3, column=3, value="TOTAL NON-EQUIPMENT COST SUBTOTAL").font = Font(bold=True)
    ws.cell(row=tr + 3, column=4, value=float(totals["non_equip"])).number_format = CURRENCY_FMT
    ws.cell(row=tr + 4, column=3, value="TOTAL CONTINGENCY").font = Font(bold=True)
    ws.cell(row=tr + 4, column=4, value=float(totals["contingency"])).number_format = CURRENCY_FMT
    ws.cell(row=tr + 5, column=3, value="TOTAL INSTALLED COST").font = Font(bold=True, size=11)
    ws.cell(row=tr + 5, column=4, value=float(totals["extended"])).number_format = CURRENCY_FMT
    ws.cell(row=tr + 5, column=4).font = Font(bold=True, size=11)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    for col in range(4, 13):
        ws.column_dimensions[get_column_letter(col)].width = 16


async def generate_bom(
    project: Project,
    systems: List[System],
    issuance_name: str,
    include_notes: bool = True,
    also_pdf: bool = False,
) -> bytes:
    """Generate a Publish BoM workbook and return as bytes."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _write_summary_sheet(wb, project, systems, issuance_name)

    for system in systems:
        _write_system_sheet(
            wb, system, project, issuance_name,
            is_bom=True,
            include_notes=include_notes,
            include_cost=True,
            include_labor_breakout=False,
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_estimate(
    project: Project,
    systems: List[System],
    issuance_name: str,
    include_notes: bool = True,
    include_cost: bool = True,
    include_labor_breakout: bool = True,
) -> bytes:
    """Generate a Publish Estimate workbook and return as bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, project, systems, issuance_name)

    for system in systems:
        _write_system_sheet(
            wb, system, project, issuance_name,
            is_bom=False,
            include_notes=include_notes,
            include_cost=include_cost,
            include_labor_breakout=include_labor_breakout,
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
