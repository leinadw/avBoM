"""Centralized equipment database CRUD."""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, func
from typing import Optional, List
import uuid

from app.db.database import get_db
from app.models.user import User
from app.models.equipment import Equipment
from app.schemas.equipment import EquipmentCreate, EquipmentUpdate, EquipmentOut
from app.auth.deps import get_current_user

router = APIRouter(prefix="/equipment", tags=["equipment"])


@router.get("/", response_model=List[EquipmentOut])
async def list_equipment(
    q: Optional[str] = Query(None, description="Search mfr, model, description"),
    category: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    stmt = select(Equipment)
    if q:
        term = f"%{q}%"
        stmt = stmt.where(
            or_(
                Equipment.mfr.ilike(term),
                Equipment.model.ilike(term),
                Equipment.description.ilike(term),
                Equipment.item_id.ilike(term),
            )
        )
    if category:
        stmt = stmt.where(Equipment.category == category)
    stmt = stmt.order_by(Equipment.mfr, Equipment.model).offset(skip).limit(limit)
    result = await db.execute(stmt)
    return result.scalars().all()


@router.post("/", response_model=EquipmentOut, status_code=201)
async def create_equipment(
    body: EquipmentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    existing = await db.execute(select(Equipment).where(Equipment.item_id == body.item_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="item_id already exists")
    equip = Equipment(**body.model_dump(), created_by_id=current_user.id)
    db.add(equip)
    await db.commit()
    await db.refresh(equip)
    return equip


@router.get("/{equipment_id}", response_model=EquipmentOut)
async def get_equipment(
    equipment_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Equipment).where(Equipment.id == equipment_id))
    equip = result.scalar_one_or_none()
    if not equip:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return equip


@router.patch("/{equipment_id}", response_model=EquipmentOut)
async def update_equipment(
    equipment_id: uuid.UUID,
    body: EquipmentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Equipment).where(Equipment.id == equipment_id))
    equip = result.scalar_one_or_none()
    if not equip:
        raise HTTPException(status_code=404, detail="Equipment not found")
    for field, val in body.model_dump(exclude_unset=True).items():
        setattr(equip, field, val)
    await db.commit()
    await db.refresh(equip)
    return equip


@router.delete("/{equipment_id}", status_code=204)
async def delete_equipment(
    equipment_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Equipment).where(Equipment.id == equipment_id))
    equip = result.scalar_one_or_none()
    if not equip:
        raise HTTPException(status_code=404, detail="Equipment not found")
    await db.delete(equip)
    await db.commit()


@router.post("/import-xlsx", response_model=List[EquipmentOut], status_code=201)
async def import_equipment_from_xlsx(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import equipment from a PROJECT_EQUIPMENT_LIST sheet in an xlsx file."""
    from openpyxl import load_workbook
    import io
    from decimal import Decimal

    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    if "PROJECT_EQUIPMENT_LIST" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="No PROJECT_EQUIPMENT_LIST sheet found")

    ws = wb["PROJECT_EQUIPMENT_LIST"]
    created = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = str(row[0]).strip() if row[0] else None
        if not item_id or item_id == "//" or item_id == "ITEM":
            continue
        mfr = str(row[1] or "").strip()
        model = str(row[2] or "").strip()
        description = str(row[3] or "").strip()
        notes = str(row[4] or "").strip()
        try:
            msrp = Decimal(str(row[5])) if row[5] else Decimal("0")
            multiplier = Decimal(str(row[6])) if row[6] else Decimal("1")
        except Exception:
            msrp = Decimal("0")
            multiplier = Decimal("1")

        existing = await db.execute(select(Equipment).where(Equipment.item_id == item_id))
        if existing.scalar_one_or_none():
            continue  # skip duplicates

        equip = Equipment(
            item_id=item_id, mfr=mfr, model=model,
            description=description, notes=notes,
            msrp=msrp, multiplier=multiplier,
            created_by_id=current_user.id,
        )
        db.add(equip)
        created.append(equip)

    await db.commit()
    for e in created:
        await db.refresh(e)
    return created
